""".. Ignore pydocstyle D400.

===================
Migrate annotations
===================

Migrate annotations from old schema to the new one. All new annotations will be
re-created from old ones.
"""
import os
from collections import defaultdict
from pathlib import Path
from typing import Dict, Iterable, Set, Union

import openpyxl

from django.core.management.base import BaseCommand
from django.db import models, transaction

from resolwe.flow.models import (
    AnnotationField,
    AnnotationGroup,
    AnnotationValue,
    Collection,
    DescriptorSchema,
    Entity,
)
from resolwe.flow.models.annotations import AnnotationType
from resolwe.flow.utils import iterate_schema


def read_mapping(file: Union[str, os.PathLike]):
    """Read the mappings file."""

    def get_row(excel_row, header):
        """Get the dictionary with header values for keys."""
        return {
            column_name: excel_row[column_index].value
            for column_index, column_name in header.items()
        }

    workbook = openpyxl.load_workbook(file)
    sheet = workbook["CHANGELOG"]

    type_map = {
        "basic:string": AnnotationType.STRING,
        "basic:decimal": AnnotationType.DECIMAL,
        "basic:integer": AnnotationType.INTEGER,
        "basic:date": AnnotationType.DATE,
    }
    header = {
        index: cell.value
        for index, cell in enumerate(next(sheet.rows))
        if cell.value is not None
    }

    # Read all the group labels in advance.
    group_labels = dict()
    for excel_row in sheet.iter_rows(2):
        row = get_row(excel_row, header)
        if row["GROUP/FIELD"] == "GROUP" and row["FROM"].strip() != "-":
            group_labels[row["NEW group name"]] = row["Label"]

    mapping = dict()
    for excel_row in sheet.iter_rows(2):
        row = get_row(excel_row, header)
        if row["GROUP/FIELD"] != "FIELD" or row["TO"].lower().strip() == "obsolete":
            continue

        choice_row = row["Choices {label : value}"]
        choices = choice_row.strip(" {}").split(", ") if choice_row else []
        vocabulary = {
            choice.split(":")[1].strip(): choice.split(":")[0].strip()
            for choice in choices
        }

        for processing_group in row["OLD group name"].split(" or "):
            field = f"{processing_group}.{row['FROM']}"
            mapping[field] = (
                row["NEW group name"],
                group_labels[row["NEW group name"]],
                row["TO"],
                row["Label"],
                row["Description"],
                type_map[row["Type"]],
                vocabulary,
            )
    return mapping


def process_descriptor_schema(
    descriptor_schema: Dict, mapping: Dict, field_map: Dict
) -> Iterable[AnnotationField]:
    """Get the annotation fields from the descriptor schema."""
    fields = set()
    for schema, _, field_full_path in iterate_schema({}, descriptor_schema):
        # Field is a dictionary contating keys name, label, type and required.
        assert field_full_path in mapping, f"{field_full_path} is not known"
        (
            group_name,
            group_label,
            field_name,
            field_label,
            field_description,
            field_type,
            field_vocabulary,
        ) = mapping[field_full_path]
        # TODO: group sort order.
        group = AnnotationGroup.objects.get_or_create(
            name=group_name, label=group_label
        )[1]
        # TODO: field sort order, field description.
        field = AnnotationField.objects.get_or_create(
            name=field_name,
            type=field_type,
            group=group,
            vocabulaly=field_vocabulary,
            label=field_label,
            description=field_description,
            validator_regex=schema.get("validate_regex"),
        )[1]
        field_map[field_full_path] = field
        fields.add(field)
    return fields


@transaction.atomic
def migrate_descriptors(mapping_file: str):
    """Create new descriptors from old ones."""
    # This should delete all the new annotation objects. Hopefully it will be fast enough.
    AnnotationGroup.objects.all().delete()
    mapping_path = Path(mapping_file)
    mapping = read_mapping(mapping_path)
    # Iterate through all descriptor schemas that are set on the Collections.
    used_descriptor_schemas = DescriptorSchema.objects.filter(
        models.Q(collection__isnull=False) | models.Q(entity__isnull=False)
    ).distinct()

    visited_fields = dict()
    for descriptor_schema in used_descriptor_schemas.iterator():
        for field, _, field_full_path in iterate_schema({}, descriptor_schema.schema):
            # Field is a dictionary contating keys name, label, type and required.
            visited_fields[field_full_path] = field

    missing_fields = set()
    for field_path in visited_fields:
        if field_path not in mapping:
            missing_fields.add(field_path)

    # Mapping between full field name and field instance.
    field_map = dict()
    # Mapping between collection and fields assigned to it.
    collection_annotation_fields: Dict[int, Set[AnnotationField]] = defaultdict(set)
    # Mapping from descriptor schema id to the list of annotation field.
    schema_fields = dict()
    # Now iterate through all the collections. Read the annotation schema of
    # its entities and create corresponding fields.
    for collection_id, schema_id, schema in (
        Entity.objects.filter(descriptor_schema__isnull=False, collection__isnull=False)
        .values_list("collection", "descriptor_schema", "descriptor_schema__schema")
        .iterator()
    ):
        if schema_id not in schema_fields:
            schema_fields[schema_id] = process_descriptor_schema(
                schema, mapping, field_map
            )
        collection_annotation_fields[collection_id].update(schema_fields[schema_id])

    for collection_id, collection_fields in collection_annotation_fields.values():
        collection: Collection = Collection.objects.get(pk=collection_id)
        collection.annotation_fields.set(collection_fields)
        # Process values for all entities in the collection.
        for entity in collection.entity_set.all():
            for _, field_value, field_full_path in iterate_schema(
                entity.descriptor, entity.descriptor_schema.schema
            ):
                field = field_map[field_full_path]
                AnnotationValue.objects.get_or_create(
                    entity=entity, field=field, value=field_value
                )


class Command(BaseCommand):
    """Migrate descriptors."""

    help = "Migrate descriptors."

    def add_arguments(self, parser):
        """Command arguments."""
        parser.add_argument(
            "mapping_file",
            help="The file with mappings from the old annotation schema to fields.",
        )

    def handle(self, *args, **options):
        """Migrate descriptors."""
        migrate_descriptors(options["mapping_file"])
